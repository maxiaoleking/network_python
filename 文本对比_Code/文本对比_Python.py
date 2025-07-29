import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# --- 文件名定义 ---
file_old = 'old.txt'
file_new = 'new.txt'
output_excel_file = 'diff_report_final.xlsx'

try:
    # --- 步骤 1: 读取文件内容 ---
    # 读取 old.txt 的所有行，保留原始顺序，作为我们报告的基准
    with open(file_old, 'r', encoding='utf-8') as f_old:
        lines_old = f_old.readlines()

    # 高效查找准备：读取 new.txt 的所有行并放入一个集合（set）中
    # 使用 .strip() 去除比对内容两端的空白字符，确保比较的准确性
    with open(file_new, 'r', encoding='utf-8') as f_new:
        lines_new_set = {line.strip() for line in f_new}

except FileNotFoundError as e:
    print(f"错误：找不到文件 '{e.filename}'。请确保脚本与 txt 文件在同一目录下。")
    exit()

# --- 步骤 2: 创建并设置 Excel 工作簿 ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "精准差异报告"

# 设置更清晰的表头
headers = ["old.txt 内容 (基准)", "在 new.txt 中是否找到?", "在 new.txt 中缺失的内容"]
ws.append(headers)
# 加粗表头
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# 定义样式
red_font = Font(color="FF0000")
not_found_fill = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")  # 淡红色背景
found_font = Font(color="008000")  # 绿色字体

# --- 步骤 3: 逐行处理并写入 Excel ---
# 以 old.txt 的行为基础进行遍历
for line in lines_old:
    stripped_line = line.strip()

    # 检查处理后的行是否存在于 new.txt 的集合中
    if stripped_line in lines_new_set:
        # 如果找到了
        status = "是"
        missing_info = ""
        row_to_write = [stripped_line, status, missing_info]
        ws.append(row_to_write)
        # 将状态单元格标为绿色
        status_cell = ws.cell(row=ws.max_row, column=2)
        status_cell.font = found_font
    else:
        # 如果没找到
        status = "否"
        missing_info = stripped_line  # 在第三列显示缺失的内容
        row_to_write = [stripped_line, status, missing_info]
        ws.append(row_to_write)
        # 将整行标记出来，方便查看
        status_cell = ws.cell(row=ws.max_row, column=2)
        status_cell.font = red_font  # 状态标红

        missing_cell = ws.cell(row=ws.max_row, column=3)
        missing_cell.font = red_font  # 缺失内容标红

        # 给整行一个淡红色背景，使其更突出
        for cell in ws[ws.max_row]:
            cell.fill = not_found_fill

# --- 步骤 4: 调整列宽并保存 ---
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 50

try:
    wb.save(output_excel_file)
    print("=" * 50)
    print(f"最终版对比完成！报告已生成在 '{output_excel_file}'。")
    print("报告说明:")
    print("1. 以 'old.txt' 为基准，每一行都进行了检查。")
    print("2. '在 new.txt 中缺失的内容' 列只在内容确实丢失时才会显示。")
    print("3. 缺失的行已用红色字体和淡红色背景高亮，方便快速定位。")
    print("=" * 50)
except Exception as e:
    print(f"保存 Excel 文件时出错: {e}")
